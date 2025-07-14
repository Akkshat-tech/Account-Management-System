#MENU PAGE- PROJECT
from tkinter import *
from tkinter import ttk

#creating base widget ( first widget)
root = Tk()
root.title("Account Management System")
root.geometry("770x440")
bg = PhotoImage(file="bg.png")
bg3 = PhotoImage(file="bg_2.png")
# f'{var}'


# opening excel
import openpyxl
from openpyxl import Workbook
import pandas as pd
wb=Workbook()
wb= openpyxl.load_workbook("database_project.xlsx")
wbcb=openpyxl.load_workbook("current_balance.xlsx")
cb=wbcb["Sheet1"]  

# canvas

mycanvas = Canvas(root,width = 770,height = 440)
mycanvas.pack()
#fill= "both",expand = True

#set image in canvas
def main():
    mycanvas.create_image(0,0,image=bg,anchor="nw")
    mycanvas.create_text(385,30,text="Account Management System",font="Helvatica 25 bold",fill="white")

#verifies the password
def verify():
    global root
    a = entry1.get()
    b = entry2.get()
    entry1.delete(0,END)
    entry2.delete(0,END)
    
    if a=="cse100" and b=="admin@123":
        mycanvas.delete('all')   #deletes everything from canvas
        main()
        mycanvas.create_text(385,100,text="User is Verified.",font="Helvatica 30 bold",fill="light blue")
        b1 = Button(root,text="Main Menu",command = menu,font="Helvatica 15 bold",activebackground="yellow")
        mycanvas.create_window(385,180,window=b1)
        
    else:
        mycanvas.delete('all')
        main()
        b2 = Button(root,text="Retry",command = login,font="Helvatica 15 bold",activebackground="yellow")
        mycanvas.create_window(385,180,window=b2) 
        mycanvas.create_text(385,100,text="User credentials do not match.",font="Helvatica 30 bold",fill="red")
                
    
def login():
    global entry1,entry2,root
    mycanvas.delete('all')  #deletes everything
    #recreating the window
    mycanvas.create_image(0,0,image=bg,anchor="nw") 
    mycanvas.create_text(385,30,text="Account Management System",font="Helvatica 25 bold",fill="white")
    
    #login credentials
    mycanvas.create_text(385,60,text="Login",font="Helvatica 30 bold",fill="yellow")
    mycanvas.create_text(80,150,text="User ID: ",font="Helvatica 20 bold",fill="white")
    entry1 = Entry(root)
    mycanvas.create_window(240,150,window=entry1)

    mycanvas.create_text(100,230,text="Password: ",font="Helvatica 20 bold",fill="white")
    entry2 = Entry(root)
    mycanvas.create_window(240,230,window=entry2)

    b_submit = Button(root,text="Submit",command=verify,activebackground="yellow")
    mycanvas.create_window(385,330,window=b_submit)    

def sure():
    def dele():
        root1.destroy()
    top = Toplevel()
    top.title("Account Management System")
    label1 = Label(top,text="Are you sure that you want to exit?",font="Helvatica 16 bold")
    label1.grid(row=0,column=1)
    ok = Button(top,text="Yes",command=dele,bg="orange",fg="white",font="Helvatica 20 bold")
    back = Button(top,text="Back",command=top.destroy,bg="cyan",fg="black",font="Helvatica 20 bold")
    ok.grid(row=1,column=1)
    back.grid(row=1,column=0)
    top.mainloop()

    
def menu():
    global root,root1
    try:
        if root.winfo_exists()==1:
            root.destroy()
        elif root2.winfo_exists()==1:
            root2.destroy()
        else:
            pass  #destroys initial widget
        try:
            if search_zero.winfo_exists()==1:
                search_zero.destroy()
        except:
            pass
    except:
        pass
    root1 = Tk()  #creates a new widget
    root1.title("Account Management System")
    bg2 = PhotoImage(file="bg_2.png")
    add = PhotoImage(file="add.png")
    graph = PhotoImage(file="graph.png")
    delete = PhotoImage(file="delete.png")
    search = PhotoImage(file="search.png")
    update = PhotoImage(file="update.png")
    mycanvas1 = Canvas(root1, width=1920, height=1080)
    mycanvas1.pack()
    mycanvas1.create_image(0,0,image=bg2,anchor="nw")
    mycanvas1.create_text(int(1920/2)-200,40,text="Account Management System",font="Helvatica 30 bold")
    mycanvas1.create_text(int(1920/2)-180,100,text="Main Menu",font="Calibri 25 bold")
    
    
    but1 = Button(root1,image=add,activebackground = "yellow",command=add_entry_feature)
    but_1 = Button(root1,text="Add Data",activebackground = "yellow",font="Calibri 25 bold",bg="brown",fg="white",command=add_entry_feature)

    but2 = Button(root1,image=search,activebackground = "yellow",command=search_entry_overall)
    but_2 = Button(root1,text="Search Data",activebackground = "yellow",command=search_entry_overall,font="Calibri 25 bold",bg="brown",fg="white")
    
    but3 = Button(root1,image=graph,command=graph_plot,activebackground = "yellow")
    but_3 = Button(root1,text="View Graphs",command=graph_plot,activebackground = "yellow",font="Calibri 25 bold",bg="brown",fg="white")

    but4 = Button(root1,image=delete,activebackground = "yellow",command=delete_entry_overall)
    but_4 = Button(root1,text="Delete Data",activebackground = "yellow",command=delete_entry_overall,font="Calibri 25 bold",bg="brown",fg="white")
    
    but5 = Button(root1,image=update,activebackground = "yellow",command=lambda: update_data(0))
    but_5 = Button(root1,text="Update data",activebackground = "yellow",command=lambda: update_data(0),font="Calibri 25 bold",bg="brown",fg="white")
    
    but6 = Button(root1,text="Exit",activebackground = "yellow",font="Calibri 25 bold",bg="brown",fg="white",command=sure)



    mycanvas1.create_window(150,250,window=but1)
    mycanvas1.create_window(150,400,window=but_1)
    mycanvas1.create_window(450,250,window=but2)
    mycanvas1.create_window(450,400,window=but_2)
    mycanvas1.create_window(750,250,window=but3)
    mycanvas1.create_window(750,400,window=but_3)
    mycanvas1.create_window(1050,250,window=but4)
    mycanvas1.create_window(1050,400,window=but_4)
    mycanvas1.create_window(1350,250,window=but5)
    mycanvas1.create_window(1350,400,window=but_5)
    mycanvas1.create_window(int(1920/2)-180,680,window=but6)
    """but2.grid(row=1,column=1)
    but3.grid(row=1,column=2)
    but4.grid(row=2,column=1)
    but5.grid(row=2,column=2,columnspan=2)
    """
    root1.mainloop()

main()
login()

def Error(field,variable):   #gives error message and destroys the error message, also creates blank at the variable
    global p_in_num
    def EXIT():
        variable.delete(0,END)
        error.destroy()
    error = Tk()
    error.title("Account Management System")
    error.geometry("440x240")
    error.title("ERROR")
    label1 = Label(error,text=str(field)).pack()
    button1 = Button(error,text = "Re-enter",command=EXIT,activebackground="yellow").pack()
    error.mainloop()


def Check_Conditions():
    global root2,mycanvas
    global p_in_num,party_name,party_address,gst_num,pan_num,mode_payment,bill_num,product,rate,quantity,total,gst,total_bill,year,month,date1
    flag = 0 #flag should be 14
    special = 0
    def char(var,name):
        satisfy = 0
        c = var.get()
        if len(c)!=0:
            satisfy=1
        else:
            Error("Please enter "+name,var)
        if satisfy==1:
            return True
        
    def integer(var,name):
        satisfy = 0
        c = var.get()
        if len(c)!=0:
            if c[0]=="-":
                p = c[1::]
                Error("Please enter valid (Non Negative) "+name,var)
                #p_in_num.delete(0,END)
            else:
                p = c
            if p.isdigit():
                satisfy=1
            else:
                Error("Please enter valid(Integer) "+name,var)
                #p_in_num.delete(0,END)
        else:
            Error("Please enter "+name,var)
            #p_in_num.delete(0,END)
            
        if satisfy==1:
            return True

    if integer(p_in_num,"Purchase invoice number"):
        flag+=1
    if char(party_name,"Party name"):
        flag+=1
    if char(party_address,"Party address"):
        flag+=1
    if char(gst_num,"GST number"):
        flag+=1
    if integer(pan_num,"PAN number"):
        flag+=1
    if char(mode_payment,"Mode of payment"):
        flag+=1
    if integer(bill_num,"Bill number"):
        flag+=1
    if char(product,"Product"):
        flag+=1
    if integer(rate,"Rate"):
        flag+=1
        special+=1
    if integer(quantity,"Quantity"):
        flag+=1
        special+=1
    if integer(gst,"GST"):
        flag+=1
        special+=1 
    if integer(year,"Year"):
        flag+=1
    if integer(month,"Month"):
        flag+=1
    if integer(date1,"Date"):
        flag+=1     #flag =14
    

    if special == 3:
        mycanvas.create_text(700,600,text="Total: ",font="Helvatica 10 bold",fill="black",anchor="w")
        mycanvas.create_text(700,700,text="Total Bill: ",font="Helvatica 10 bold",fill="black",anchor="w")
    
        total_value = int(rate.get())*int(quantity.get())
        total = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
        total.insert(0,str(total_value))
        total.config(state=DISABLED)

        amount=(int(gst.get())/100)*total_value+total_value
        total_bill = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
        total_bill.insert(0,str(amount))
        total_bill.config(state=DISABLED)
        mycanvas.create_window(900,600,window=total)
        mycanvas.create_window(900,700,window=total_bill)

    if flag==14 and special==3:
        def back():
            global root2,mycanvas
            global p_in_num,party_name,party_address,gst_num,pan_num,mode_payment,bill_num,product,rate,quantity,total,gst,total_bill,year,month,date1
    
            p_in_num.config(state=NORMAL)
            party_name.config(state=NORMAL)
            party_address.config(state=NORMAL)
            gst_num.config(state=NORMAL)
            pan_num.config(state=NORMAL)
            mode_payment.config(state=NORMAL)
            bill_num.config(state=NORMAL)
            product.config(state=NORMAL)
            rate.config(state=NORMAL)
            quantity.config(state=NORMAL)
            gst.config(state=NORMAL)
            year.config(state=NORMAL)
            month.config(state=NORMAL)
            date1.config(state=NORMAL)
            top.destroy()

        def mssg():
            ws1=wb["purchase invoice"]
            global p_in_num,party_name,party_address,gst_num,pan_num,mode_payment,bill_num,product,rate,quantity,total,gst,total_bill,year,month,date1
            list1 = [int(p_in_num.get()),int(month.get()),int(date1.get()),int(year.get()),party_name.get(),party_address.get(),gst_num.get(),int(pan_num.get()),(mode_payment.get()).upper(),int(bill_num.get()),product.get(),int(rate.get()),int(quantity.get()),int(total.get()),int(gst.get()),float(total_bill.get())]
                
            def ok():
                top2.destroy()
                root2.destroy()
                menu()
            ws1.append(list1)
            money=cb["A1"]
            cb["A1"]=cb["A1"].value+list1[-1]
            wb.save("database_project.xlsx")
            wbcb.save("current_balance.xlsx")
            top.destroy()
            top2=Tk()
            top2.title("Account Management System")
            label = Label(top2,text="Data added to database.",font="Helvatica 15 bold",fg='green').grid(row=0,column=0,columnspan=2)
            confirm = Button(top2,text="Ok",bg="cyan",command=ok,fg="white",font="Calibri 10 bold").grid(row=1,column=0)
            
            
        p_in_num.config(state=DISABLED)
        party_name.config(state=DISABLED)
        party_address.config(state=DISABLED)
        gst_num.config(state=DISABLED)
        pan_num.config(state=DISABLED)
        mode_payment.config(state=DISABLED)
        bill_num.config(state=DISABLED)
        product.config(state=DISABLED)
        rate.config(state=DISABLED)
        quantity.config(state=DISABLED)
        gst.config(state=DISABLED)
        year.config(state=DISABLED)
        month.config(state=DISABLED)
        date1.config(state=DISABLED)
        top = Tk()
        top.title("Account Management System")
        label = Label(top,text="Kindly check all the entries before adding to database.",font="Helvatica 15 bold").grid(row=0,column=0,columnspan=2)
        confirm = Button(top,text="Confirm",bg="cyan",command=mssg,fg="white",font="Calibri 10 bold").grid(row=1,column=0)
        back = Button(top,text="Back",bg="cyan",fg="white",command=back,font="Calibri 10 bold").grid(row=1,column=1)


#sales invoice


def Check_Conditions2():
    global root2,mycanvas
    root2.title("Account Management System")
    global s_in_num,party_name,party_address,gst_num,pan_num,mode_payment,bill_num,product,rate,quantity,total,gst,total_bill,year,month,date1
    flag = 0 #flag should be 14
    special = 0
    def char(var,name):
        satisfy = 0
        c = var.get()
        if len(c)!=0:
            satisfy=1
        else:
            Error("Please enter "+name,var)
        if satisfy==1:
            return True
        
    def integer(var,name):
        satisfy = 0
        c = var.get()
        if len(c)!=0:
            if c[0]=="-":
                p = c[1::]
                Error("Please enter valid (Non Negative) "+name,var)
                #p_in_num.delete(0,END)
            else:
                p = c
            if p.isdigit():
                satisfy=1
            else:
                Error("Please enter valid(Integer) "+name,var)
                #p_in_num.delete(0,END)
        else:
            Error("Please enter "+name,var)
            #p_in_num.delete(0,END)
            
        if satisfy==1:
            return True

    if integer(s_in_num,"Sales invoice number"):
        flag+=1
    if char(party_name,"Party name"):
        flag+=1
    if char(party_address,"Party address"):
        flag+=1
    if char(gst_num,"GST number"):
        flag+=1
    if integer(pan_num,"PAN number"):
        flag+=1
    if char(mode_payment,"Mode of payment"):
        flag+=1
    if integer(bill_num,"Bill number"):
        flag+=1
    if char(product,"Product"):
        flag+=1
    if integer(rate,"Rate"):
        flag+=1
        special+=1
    if integer(quantity,"Quantity"):
        flag+=1
        special+=1
    if integer(gst,"GST"):
        flag+=1
        special+=1 
    if integer(year,"Year"):
        flag+=1
    if integer(month,"Month"):
        flag+=1
    if integer(date1,"Date"):
        flag+=1     #flag =14
    

    if special == 3:
        mycanvas.create_text(700,600,text="Total: ",font="Helvatica 10 bold",fill="black",anchor="w")
        mycanvas.create_text(700,700,text="Total Bill: ",font="Helvatica 10 bold",fill="black",anchor="w")
    
        total_value = int(rate.get())*int(quantity.get())
        total = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
        total.insert(0,str(total_value))
        total.config(state=DISABLED)

        amount=(int(gst.get())/100)*total_value+total_value
        total_bill = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
        total_bill.insert(0,str(amount))
        total_bill.config(state=DISABLED)
        mycanvas.create_window(900,600,window=total)
        mycanvas.create_window(900,700,window=total_bill)

    if flag==14 and special==3:
        def back():
            global root2,mycanvas
            global s_in_num,party_name,party_address,gst_num,pan_num,mode_payment,bill_num,product,rate,quantity,total,gst,total_bill,year,month,date1
    
            s_in_num.config(state=NORMAL)
            party_name.config(state=NORMAL)
            party_address.config(state=NORMAL)
            gst_num.config(state=NORMAL)
            pan_num.config(state=NORMAL)
            mode_payment.config(state=NORMAL)
            bill_num.config(state=NORMAL)
            product.config(state=NORMAL)
            rate.config(state=NORMAL)
            quantity.config(state=NORMAL)
            gst.config(state=NORMAL)
            year.config(state=NORMAL)
            month.config(state=NORMAL)
            date1.config(state=NORMAL)
            top.destroy()

        def mssg():
            ws1=wb["sales invoice"]
            global s_in_num,party_name,party_address,gst_num,pan_num,mode_payment,bill_num,product,rate,quantity,total,gst,total_bill,year,month,date1
            list1 = [int(s_in_num.get()),int(month.get()),int(date1.get()),int(year.get()),party_name.get(),party_address.get(),gst_num.get(),int(pan_num.get()),(mode_payment.get()).upper(),int(bill_num.get()),product.get(),int(rate.get()),int(quantity.get()),int(total.get()),int(gst.get()),float(total_bill.get())]
                
            def ok():
                top2.destroy()
                root2.destroy()
                menu()
            ws1.append(list1)
            money=cb["A1"]
            cb["A1"]=cb["A1"].value+list1[-1]
            wb.save("database_project.xlsx")
            wbcb.save("current_balance.xlsx")
            top.destroy()
            top2=Tk()
            top2.title("Account Management System")
            label = Label(top2,text="Data added to database.",font="Helvatica 15 bold",fg='green').grid(row=0,column=0,columnspan=2)
            confirm = Button(top2,text="Ok",bg="cyan",command=ok,fg="white",font="Calibri 10 bold").grid(row=1,column=0)
            
            
        s_in_num.config(state=DISABLED)
        party_name.config(state=DISABLED)
        party_address.config(state=DISABLED)
        gst_num.config(state=DISABLED)
        pan_num.config(state=DISABLED)
        mode_payment.config(state=DISABLED)
        bill_num.config(state=DISABLED)
        product.config(state=DISABLED)
        rate.config(state=DISABLED)
        quantity.config(state=DISABLED)
        gst.config(state=DISABLED)
        year.config(state=DISABLED)
        month.config(state=DISABLED)
        date1.config(state=DISABLED)
        top = Tk()
        top.title("Account Management System")
        label = Label(top,text="Kindly check all the entries before adding to database.",font="Helvatica 15 bold").grid(row=0,column=0,columnspan=2)
        confirm = Button(top,text="Confirm",bg="cyan",command=mssg,fg="white",font="Calibri 10 bold").grid(row=1,column=0)
        back = Button(top,text="Back",bg="cyan",fg="white",command=back,font="Calibri 10 bold").grid(row=1,column=1)
            
# purchase invoice            

def add_data():
    global p_in_num,party_name,party_address,gst_num,pan_num,mode_payment,bill_num,product,rate,quantity,total,gst,total_bill,year,month,date
    global root2,date1,mycanvas
    # creates new widget and sets background
    root2.destroy()
    root2 = Tk()
    root2.title("Account Management System")
    bg2 = PhotoImage(file="bg_2.png")
    #creates buttons and canvas    
    mycanvas = Canvas(root2,width=1920,height=1080)
    mycanvas.pack() #creating a new canvas
    mycanvas.create_image(0,0,image=bg2,anchor="nw")  #sets background of mycanvas
    mycanvas.create_text(int(1920/2)-200,30,text="Account Management System",font="Helvatica 30 bold",fill="black")
    mycanvas.create_text(300,100,text="Kindly fill the following details: ",font="Helvatica 25 bold",fill="black")

    
    '''def date(e):
        global date1
        l2 = [] #31 days
        l3 = [] #29 days
        l4 = [] #28 days
        
        for i in range(1,32):
            l2.append(i)
            if i<30:
                l3.append(i)
            elif(i<29):
                l4.append(i)    

        if (int(year.get())%4!=0):
            if month.get()=="February":
                date1.config(value=l4)
            else:
                date1.config(value=l2)
            
            
        if (int(year.get())%4==0):
            if (int(year.get())%100==0):
                if (int(year.get())%400==0):
                    if month.get()=="February":
                        date1.config(value=l3)
                    else:
                        date1.config(value=l2)
                else:
                    if month.get()=="February":
                        date1.config(value=l4)
                    else:
                        date1.config(value=l2)
            else:
                if month.get()=="February":
                    date1.config(value=l3)
                else:
                    date1.config(value=l2)'''
       
                     
    # creating text input options
    mycanvas.create_text(100,200,text="Purchase invoice number: ",font="Helvatica 10 bold",fill="black",anchor="w")
    mycanvas.create_text(100,250,text="(P_in_num)Must be greater than 0",font="Helvatica 10 bold",fill="black",anchor="w")
    
    mycanvas.create_text(100,300,text="Party Name: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(100,400,text="Party Address: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(100,500,text="GST No:",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(100,600,text="PAN No:",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(100,700,text="Mode of Payment: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(1150,200,text="Bill No: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(1150,300,text="Year: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(1150,400,text="Month: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(1150,500,text="Day: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(700,200,text="Product: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(700,300,text="Rate: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(700,400,text="Quantity: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(700,500,text="GST: ",font="Helvatica 15 bold",fill="black",anchor="w")
    #Creating dropdown list box for date entries
    #months = ["January","February","March","April","May","June","July","August","September","October","November","December"]

    l1 = [1,2,3,4,5,6,7,8,9,10,11,12]
    year_list = []
    for i in range(1970,2025):
        year_list.append(i)
    
    year = ttk.Combobox(root2,value=year_list)
    year.current(int(2024-1970))
    mycanvas.create_window(1300,300,window=year)
    #year.bind("<<ComboboxSelected>>",date)

    month = ttk.Combobox(root2,value=l1)
    month.current(0)
    mycanvas.create_window(1300,400,window=month)
    #month.bind("<<ComboboxSelected>>",date)

    day=[]
    for i in range(1,32):
            day.append(i)
    date1 = ttk.Combobox(root2,value=day)
    mycanvas.create_window(1300,500,window=date1)
    #date1.bind("<<ComboboxSelected>>",date)
    
    # create entries
    p_in_num = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    party_name = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    party_address = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    gst_num = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    pan_num = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    mode_payment = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    bill_num = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    product = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    rate = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    quantity = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    #total = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold",state=DISABLED)
    gst = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    #total_bill = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    #total_bill.insert(0,"3456")
    #total_bill.config(state=DISABLED)
    # put on canvas
    mycanvas.create_window(400,200,window=p_in_num)
    mycanvas.create_window(400,300,window=party_name)
    mycanvas.create_window(400,400,window=party_address)
    mycanvas.create_window(400,500,window=gst_num)
    mycanvas.create_window(400,600,window=pan_num)
    mycanvas.create_window(400,700,window=mode_payment)
    mycanvas.create_window(1350,200,window=bill_num)
    mycanvas.create_window(900,200,window=product)
    mycanvas.create_window(900,300,window=rate)
    mycanvas.create_window(900,400,window=quantity)
    #mycanvas.create_window(900,500,window=total)
    mycanvas.create_window(900,500,window=gst)
    #mycanvas.create_window(900,700,window=total_bill)

    def delete():
        p_in_num.delete(0,END)
        party_name.delete(0,END)
        party_address.delete(0,END)
        gst_num.delete(0,END)
        pan_num.delete(0,END)
        mode_payment.delete(0,END)
        bill_num.delete(0,END)
        product.delete(0,END)
        rate.delete(0,END)
        quantity.delete(0,END)
        total.delete(0,END)
        gst.delete(0,END)
        total_bill.delete(0,END)
    def revisit_back():
        try:
            root2.destroy()
        except:
            pass
        add_entry_feature()

    def sure2():
        top = Toplevel()
        label = Label(top,text="Your progress will be lost if you go back").grid(row=0,column=0,columnspan=2)
        but1 = Button(top,text="Stay here",fg="black",activebackground="yellow",bg="cyan",font="Calibri 15 bold",command=top.destroy).grid(row=1,column=0)
        but2 = Button(top,text="Back",fg="black",bg="grey",activebackground="red",font="Calibri 15 bold",command=revisit_back).grid(row=1,column=1)

        
    clear_all = Button(root2,text="Clear all the data",activebackground="yellow",command=delete,font="Calibri 15 bold")
    mycanvas.create_window(int(1920/2)-100,750,window=clear_all)
    #back
    back2 = Button(root2,text="Back",activebackground = "yellow",font="Calibri 15 bold",bg="brown",fg="white",command=sure2)
    mycanvas.create_window(50,30,window=back2)
    # check button:
    check=Button(root2,text="Check",bg="brown",fg="yellow",font="Helvatica 20 bold",command=Check_Conditions,activebackground="yellow")
    mycanvas.create_window(int(1920/2)-300,750,window=check)
    root2.mainloop()





# for sales invoice

def add_data2():
    global s_in_num,party_name,party_address,gst_num,pan_num,mode_payment,bill_num,product,rate,quantity,total,gst,total_bill,year,month,date
    global root2,date1,mycanvas
    # creates new widget and sets background
    root2.destroy()
    root2 = Tk()
    root2.title("Account Management System")
    bg2 = PhotoImage(file="bg_2.png")
    #creates buttons and canvas    
    mycanvas = Canvas(root2,width=1920,height=1080)
    mycanvas.pack() #creating a new canvas
    mycanvas.create_image(0,0,image=bg2,anchor="nw")  #sets background of mycanvas
    mycanvas.create_text(int(1920/2)-200,30,text="Account Management System",font="Helvatica 30 bold",fill="black")
    mycanvas.create_text(300,100,text="Kindly fill the following details: ",font="Helvatica 25 bold",fill="black")

    
    '''def date(e):
        global date1
        l2 = [] #31 days
        l3 = [] #29 days
        l4 = [] #28 days
        
        for i in range(1,32):
            l2.append(i)
            if i<30:
                l3.append(i)
            elif(i<29):
                l4.append(i)    

        if (int(year.get())%4!=0):
            if month.get()=="February":
                date1.config(value=l4)
            else:
                date1.config(value=l2)
            
            
        if (int(year.get())%4==0):
            if (int(year.get())%100==0):
                if (int(year.get())%400==0):
                    if month.get()=="February":
                        date1.config(value=l3)
                    else:
                        date1.config(value=l2)
                else:
                    if month.get()=="February":
                        date1.config(value=l4)
                    else:
                        date1.config(value=l2)
            else:
                if month.get()=="February":
                    date1.config(value=l3)
                else:
                    date1.config(value=l2)'''
       
                     
    # creating text input options
    mycanvas.create_text(100,200,text="Sales invoice number: ",font="Helvatica 10 bold",fill="black",anchor="w")
    mycanvas.create_text(100,250,text="(S_in_num)Must be greater than 0",font="Helvatica 10 bold",fill="black",anchor="w")
    
    mycanvas.create_text(100,300,text="Party Name: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(100,400,text="Party Address: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(100,500,text="GST No:",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(100,600,text="PAN No:",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(100,700,text="Mode of Payment: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(1150,200,text="Bill No: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(1150,300,text="Year: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(1150,400,text="Month: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(1150,500,text="Day: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(700,200,text="Product: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(700,300,text="Rate: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(700,400,text="Quantity: ",font="Helvatica 15 bold",fill="black",anchor="w")
    mycanvas.create_text(700,500,text="GST: ",font="Helvatica 15 bold",fill="black",anchor="w")
    #Creating dropdown list box for date entries
    #months = ["January","February","March","April","May","June","July","August","September","October","November","December"]

    l1 = [1,2,3,4,5,6,7,8,9,10,11,12]
    year_list = []
    for i in range(1970,2025):
        year_list.append(i)
    
    year = ttk.Combobox(root2,value=year_list)
    year.current(int(2024-1970))
    mycanvas.create_window(1300,300,window=year)
    #year.bind("<<ComboboxSelected>>",date)

    month = ttk.Combobox(root2,value=l1)
    month.current(0)
    mycanvas.create_window(1300,400,window=month)
    #month.bind("<<ComboboxSelected>>",date)

    day=[]
    for i in range(1,32):
            day.append(i)
    date1 = ttk.Combobox(root2,value=day)
    mycanvas.create_window(1300,500,window=date1)
    #date1.bind("<<ComboboxSelected>>",date)
    
    # create entries
    s_in_num = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    party_name = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    party_address = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    gst_num = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    pan_num = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    mode_payment = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    bill_num = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    product = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    rate = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    quantity = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    #total = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold",state=DISABLED)
    gst = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    #total_bill = Entry(root2,bg="brown",fg="white",font="Helvatica 13 bold")
    #total_bill.insert(0,"3456")
    #total_bill.config(state=DISABLED)
    # put on canvas
    mycanvas.create_window(400,200,window=s_in_num)
    mycanvas.create_window(400,300,window=party_name)
    mycanvas.create_window(400,400,window=party_address)
    mycanvas.create_window(400,500,window=gst_num)
    mycanvas.create_window(400,600,window=pan_num)
    mycanvas.create_window(400,700,window=mode_payment)
    mycanvas.create_window(1350,200,window=bill_num)
    mycanvas.create_window(900,200,window=product)
    mycanvas.create_window(900,300,window=rate)
    mycanvas.create_window(900,400,window=quantity)
    #mycanvas.create_window(900,500,window=total)
    mycanvas.create_window(900,500,window=gst)
    #mycanvas.create_window(900,700,window=total_bill)

    def delete():
        s_in_num.delete(0,END)
        party_name.delete(0,END)
        party_address.delete(0,END)
        gst_num.delete(0,END)
        pan_num.delete(0,END)
        mode_payment.delete(0,END)
        bill_num.delete(0,END)
        product.delete(0,END)
        rate.delete(0,END)
        quantity.delete(0,END)
        total.delete(0,END)
        gst.delete(0,END)
        total_bill.delete(0,END)
    def revisit_back():
        try:
            root2.destroy()
        except:
            pass
        add_entry_feature()

    def sure2():
        top = Toplevel()
        top.title("Account Management System")
        label = Label(top,text="Your progress will be lost if you go back").grid(row=0,column=0,columnspan=2)
        but1 = Button(top,text="Stay here",fg="black",activebackground="yellow",bg="cyan",font="Calibri 15 bold",command=top.destroy).grid(row=1,column=0)
        but2 = Button(top,text="Back",fg="black",bg="grey",activebackground="red",font="Calibri 15 bold",command=revisit_back).grid(row=1,column=1)

        
    clear_all = Button(root2,text="Clear all the data",activebackground="yellow",command=delete,font="Calibri 15 bold")
    mycanvas.create_window(int(1920/2)-100,750,window=clear_all)
    #back
    back2 = Button(root2,text="Back",activebackground = "yellow",font="Calibri 15 bold",bg="brown",fg="white",command=sure2)
    mycanvas.create_window(50,30,window=back2)
    # check button:
    check=Button(root2,text="Check",bg="brown",fg="yellow",font="Helvatica 20 bold",command=Check_Conditions2,activebackground="yellow")
    mycanvas.create_window(int(1920/2)-300,750,window=check)
    root2.mainloop()    

def add_entry_feature():
    global root1,root2 #declares global variable
    def revisit():
        root2.destroy()
        menu()
    try:
        root1.destroy()#destroys previous widget
    except:
        pass
    root2 = Tk()
    root2.title("Account Management System")
    bg2 = PhotoImage(file="bg_2.png")
        
    mycanvas = Canvas(root2,width=1920,height=1080)
    mycanvas.pack() #creating a new canvas
    mycanvas.create_image(0,0,image=bg2,anchor="nw")  #sets background of mycanvas
    mycanvas.create_text(int(1920/2)-200,30,text="Account Management System",font="Helvatica 25 bold",fill="black")

    button1 = Button(root2,text="Add entry to Purchase Invoice",font="Calibri 30",bg="brown",fg="white",command=add_data,activebackground="yellow")
    mycanvas.create_window(int(1920/2)-200,250,window=button1)

    button2 = Button(root2,text="Add entry to Sales Invoice",font="Calibri 30",bg="brown",fg="white",command=add_data2,activebackground="yellow")
    mycanvas.create_window(int(1920/2)-200,500,window=button2)

    
    back = Button(root2,text="Back",font="Calibri 15",bg="brown",fg="white",command=revisit,activebackground="yellow")
    mycanvas.create_window(50,30,window=back)

    
    root2.mainloop()

#Purchase invoice

def search1():
    sheet=wb["purchase invoice"]
    global bg3,search_zero,enter,root,mycanvas
    def again():
        root.destroy()
        search_entry_overall()
    search_zero.destroy()
    root = Tk()
    root.title("Account Management System")
    #set background
    bg3 = PhotoImage(file="bg_2.png")
    mycanvas = Canvas(root,width=1920,height=1080)
    mycanvas.pack()
    mycanvas.create_image(0,0,image=bg3,anchor="nw")
    mycanvas.create_text(int(1920/2)-200,40,text="Account Management System",font="Helvatica 30 bold")


    # create search option
    style = ttk.Style()
    style.configure("TCombobox",fieldbackground="brown",background="white")
    
    mycanvas.create_text(200,118,text="Search by: ",font="Calibri 20 bold")
    option_list = ["Invoice number","Party name","Date","Month","Year"]
    options = ttk.Combobox(root,value=option_list,font="Helvatica 15 bold")
    options.current(0)
    mycanvas.create_window(400,120,window=options)

    def data():
        global c,enter,num
        enter.config(state=DISABLED)

        def lock_goto(var,num):
            var.config(state=DISABLED)
            purchase = var.get()
            sheet=wb["purchase invoice"] 
            search(0,num,purchase,sheet)
        
        c = options.get()    

        if c=="Invoice number":
            mycanvas.create_text(200,250,text="Enter the invoice number: ",font="Calibri 20 bold")
            in_num = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(470,255,window=in_num)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(in_num,1),activebackground="yellow")
            mycanvas.create_window(630,250,window=ser)
            
        if c=="Party name":
            mycanvas.create_text(200,250,text="Enter the Party name: ",font="Calibri 20 bold")
            p_name = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(470,255,window=p_name)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(p_name,2),activebackground="yellow")
            mycanvas.create_window(630,250,window=ser)
            
        if c=="Date":
            mycanvas.create_text(200,250,text="Enter the Date (DD): ",font="Calibri 20 bold")
            date = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(440,255,window=date)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(date,3),activebackground="yellow")
            mycanvas.create_window(600,250,window=ser)
            
        if c=="Month":
            mycanvas.create_text(200,250,text="Enter the Month: ",font="Calibri 20 bold")
            month = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(450,255,window=month)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(month,4),activebackground="yellow")
            mycanvas.create_window(610,250,window=ser)
            
        if c=="Year":
            mycanvas.create_text(200,250,text="Enter the Year: ",font="Calibri 20 bold")
            year = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(440,255,window=year)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(year,5),activebackground="yellow")
            mycanvas.create_window(600,250,window=ser)
            
        

    enter = Button(root,text="Continue",font="Calibri 15 bold",command=data,bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(400,180,window=enter)

    
    
    back3 = Button(root,text="Back",font="Calibri 15 bold",command=again,bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(50,30,window=back3)
    root.mainloop()

#Sales invoice

def search2():
    sheet=wb["sales invoice"]
    global bg3,search_zero,enter,root,mycanvas
    def again():
        root.destroy()
        search_entry_overall()
    search_zero.destroy()
    root = Tk()
    root.title("Account Management System")
    #set background
    bg3 = PhotoImage(file="bg_2.png")
    mycanvas = Canvas(root,width=1920,height=1080)
    mycanvas.pack()
    mycanvas.create_image(0,0,image=bg3,anchor="nw")
    mycanvas.create_text(int(1920/2)-200,40,text="Account Management System",font="Helvatica 30 bold")


    # create search option
    style = ttk.Style()
    style.configure("TCombobox",fieldbackground="brown",background="white")
    
    mycanvas.create_text(200,118,text="Search by: ",font="Calibri 20 bold")
    option_list = ["Invoice number","Party name","Date","Month","Year"]
    options = ttk.Combobox(root,value=option_list,font="Helvatica 15 bold")
    options.current(0)
    mycanvas.create_window(400,120,window=options)

    def data():
        global c,enter,num
        enter.config(state=DISABLED)

        def lock_goto(var,num):
            var.config(state=DISABLED)
            purchase = var.get()
            sheet=wb["sales invoice"] 
            search_2(0,num,purchase,sheet)
        
        c = options.get()    

        if c=="Invoice number":
            mycanvas.create_text(200,250,text="Enter the invoice number: ",font="Calibri 20 bold")
            in_num = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(470,255,window=in_num)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(in_num,1),activebackground="yellow")
            mycanvas.create_window(630,250,window=ser)
            
        if c=="Party name":
            mycanvas.create_text(200,250,text="Enter the Party name: ",font="Calibri 20 bold")
            p_name = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(470,255,window=p_name)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(p_name,2),activebackground="yellow")
            mycanvas.create_window(630,250,window=ser)
            
        if c=="Date":
            mycanvas.create_text(200,250,text="Enter the Date (DD): ",font="Calibri 20 bold")
            date = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(440,255,window=date)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(date,3),activebackground="yellow")
            mycanvas.create_window(600,250,window=ser)
            
        if c=="Month":
            mycanvas.create_text(200,250,text="Enter the Month: ",font="Calibri 20 bold")
            month = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(450,255,window=month)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(month,4),activebackground="yellow")
            mycanvas.create_window(610,250,window=ser)
            
        if c=="Year":
            mycanvas.create_text(200,250,text="Enter the Year: ",font="Calibri 20 bold")
            year = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(440,255,window=year)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(year,5),activebackground="yellow")
            mycanvas.create_window(600,250,window=ser)
            
        

    enter = Button(root,text="Continue",font="Calibri 15 bold",command=data,bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(400,180,window=enter)

    
    
    back3 = Button(root,text="Back",font="Calibri 15 bold",command=again,bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(50,30,window=back3)
    root.mainloop()




# Search _ frontend

def search_entry_overall():

    global root1,search_zero
    global bg3
    #def revisit_menu2():
    def again():
        try:
            search_zero.destroy()
            menu()
        except:
            pass
        
    try:
        root1.destroy()
    except:
        pass
    search_zero = Tk()
    search_zero.title("Account Management System")
    mycanvas = Canvas(search_zero,width=1920,height=1080)
    mycanvas.pack()
    #set background
    bg3 = PhotoImage(file="bg_2.png")
    mycanvas.create_image(0,0,image=bg3,anchor="nw")
    mycanvas.create_text(int(1920/2)-200,40,text="Account Management System",font="Helvatica 30 bold")

    #add other options
    button1 = Button(search_zero,text="Search in Purchase Invoice",command=search1,font="Calibri 30",bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(int(1920/2)-200,250,window=button1)

    button2 = Button(search_zero,text="Search in Sales Invoice",command=search2,font="Calibri 30",bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(int(1920/2)-200,550,window=button2)

    back3 = Button(search_zero,text="Back",font="Calibri 15 bold",command=again,bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(50,30,window=back3)
    search_zero.mainloop()

#sales invoice

def search_2(n,num,purchase,sheet):
    found_list = []
    found = False
    if (num == 1):

        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    found = True
                    found_list.append(row_values)
            n = n+1
        
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Sales invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Sales invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Sales invoice number",text = "Sales invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)
        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')
        
    elif (num == 2):

        for row in sheet.iter_rows(min_row=1, min_col=5, max_row=n, max_col=5):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    found_list.append(row_values)
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Sales invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Sales invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Sales invoice number",text = "Sales invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)
        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

            
    elif (num == 3):

        for row in sheet.iter_rows(min_row=1, min_col=3, max_row=n, max_col=3):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    found_list.append(row_values)
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Sales invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Sales invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Sales invoice number",text = "Sales invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)
        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

    elif (num == 4):

        for row in sheet.iter_rows(min_row=1, min_col=2, max_row=n, max_col=2):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    found_list.append(row_values)
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Sales invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Sales invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Sales invoice number",text = "Sales invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)
        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

    elif (num == 5):

        for row in sheet.iter_rows(min_row=1, min_col=4, max_row=n, max_col=4):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    found_list.append(row_values)
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Sales invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Sales invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Sales invoice number",text = "Sales invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)
        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

    else:
        print("Enter a number from the list only")

#Purchase invoice

def search(n,num,purchase,sheet):
    found_list = []
    found = False
    if (num == 1):

        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    found_list.append(row_values)
            n = n+1
        
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Purchase invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Purchase invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Purchase invoice number",text = "Purchase invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)
        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')
        
    elif (num == 2):

        for row in sheet.iter_rows(min_row=1, min_col=5, max_row=n, max_col=5):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    found_list.append(row_values)
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Purchase invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Purchase invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Purchase invoice number",text = "Purchase invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)
        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

            
    elif (num == 3):

        for row in sheet.iter_rows(min_row=1, min_col=3, max_row=n, max_col=3):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    found_list.append(row_values)
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Purchase invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Purchase invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Purchase invoice number",text = "Purchase invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)
        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

    elif (num == 4):

        for row in sheet.iter_rows(min_row=1, min_col=2, max_row=n, max_col=2):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    found_list.append(row_values)
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Purchase invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Purchase invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Purchase invoice number",text = "Purchase invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)
        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

    elif (num == 5):

        for row in sheet.iter_rows(min_row=1, min_col=4, max_row=n, max_col=4):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    found_list.append(row_values)
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Purchase invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Purchase invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Purchase invoice number",text = "Purchase invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)
        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

    else:
        print("Enter a number from the list only")



def delete1():
    sheet=wb["purchase invoice"]
    global bg3,delete_zero,enter,root,mycanvas
    def again():
        root.destroy()
        delete_entry_overall()
    search_zero.destroy()
    root = Tk()
    root.title("Account Management System")
    #set background
    bg3 = PhotoImage(file="bg_2.png")
    mycanvas = Canvas(root,width=1920,height=1080)
    mycanvas.pack()
    mycanvas.create_image(0,0,image=bg3,anchor="nw")
    mycanvas.create_text(int(1920/2)-200,40,text="Account Management System",font="Helvatica 30 bold")


    # create search option
    style = ttk.Style()
    style.configure("TCombobox",fieldbackground="brown",background="white")
    
    mycanvas.create_text(200,118,text="Delete: ",font="Calibri 20 bold")
    option_list = ["Invoice number","Party name","Date","Month","Year"]
    options = ttk.Combobox(root,value=option_list,font="Helvatica 15 bold")
    options.current(0)
    mycanvas.create_window(400,120,window=options)

    def data():
        global c,enter,num
        enter.config(state=DISABLED)

        def lock_goto(var,num):
            ser.config(state=DISABLED)
            var.config(state=DISABLED)
            purchase = var.get()
            sheet=wb["purchase invoice"] 
            delete(0,num,purchase,sheet)
        
        c = options.get()    

        if c=="Invoice number":
            mycanvas.create_text(200,250,text="Enter the invoice number: ",font="Calibri 20 bold")
            in_num = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(470,255,window=in_num)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(in_num,1),activebackground="yellow")
            mycanvas.create_window(630,250,window=ser)
            
        if c=="Party name":
            mycanvas.create_text(200,250,text="Enter the Party name: ",font="Calibri 20 bold")
            p_name = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(470,255,window=p_name)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(p_name,2),activebackground="yellow")
            mycanvas.create_window(630,250,window=ser)
            
        if c=="Date":
            mycanvas.create_text(200,250,text="Enter the Date (DD): ",font="Calibri 20 bold")
            date = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(440,255,window=date)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(date,3),activebackground="yellow")
            mycanvas.create_window(600,250,window=ser)
            
        if c=="Month":
            mycanvas.create_text(200,250,text="Enter the Month: ",font="Calibri 20 bold")
            month = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(450,255,window=month)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(month,4),activebackground="yellow")
            mycanvas.create_window(610,250,window=ser)
            
        if c=="Year":
            mycanvas.create_text(200,250,text="Enter the Year: ",font="Calibri 20 bold")
            year = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(440,255,window=year)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(year,5),activebackground="yellow")
            mycanvas.create_window(600,250,window=ser)
            
        
    enter = Button(root,text="Continue",font="Calibri 15 bold",command=data,bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(400,180,window=enter)

    
    
    back3 = Button(root,text="Back",font="Calibri 15 bold",command=again,bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(50,30,window=back3)
    root.mainloop()

#Sales Invoice

def delete2():
    sheet=wb["sales invoice"]
    global bg3,delete_zero,enter,root,mycanvas
    def again():
        root.destroy()
        delete_entry_overall()
    search_zero.destroy()
    root = Tk()
    root.title("Account Management System")
    #set background
    bg3 = PhotoImage(file="bg_2.png")
    mycanvas = Canvas(root,width=1920,height=1080)
    mycanvas.pack()
    mycanvas.create_image(0,0,image=bg3,anchor="nw")
    mycanvas.create_text(int(1920/2)-200,40,text="Account Management System",font="Helvatica 30 bold")


    # create search option
    style = ttk.Style()
    style.configure("TCombobox",fieldbackground="brown",background="white")
    
    mycanvas.create_text(200,118,text="Delete: ",font="Calibri 20 bold")
    option_list = ["Invoice number","Party name","Date","Month","Year"]
    options = ttk.Combobox(root,value=option_list,font="Helvatica 15 bold")
    options.current(0)
    mycanvas.create_window(400,120,window=options)

    def data():
        global c,enter,num
        enter.config(state=DISABLED)

        def lock_goto(var,num):
            ser.config(state=DISABLED)
            var.config(state=DISABLED)
            purchase = var.get()
            sheet=wb["sales invoice"] 
            delete_2(0,num,purchase,sheet)
        
        c = options.get()    

        if c=="Invoice number":
            mycanvas.create_text(200,250,text="Enter the invoice number: ",font="Calibri 20 bold")
            in_num = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(470,255,window=in_num)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(in_num,1),activebackground="yellow")
            mycanvas.create_window(630,250,window=ser)
            
        if c=="Party name":
            mycanvas.create_text(200,250,text="Enter the Party name: ",font="Calibri 20 bold")
            p_name = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(470,255,window=p_name)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(p_name,2),activebackground="yellow")
            mycanvas.create_window(630,250,window=ser)
            
        if c=="Date":
            mycanvas.create_text(200,250,text="Enter the Date (DD): ",font="Calibri 20 bold")
            date = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(440,255,window=date)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(date,3),activebackground="yellow")
            mycanvas.create_window(600,250,window=ser)
            
        if c=="Month":
            mycanvas.create_text(200,250,text="Enter the Month: ",font="Calibri 20 bold")
            month = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(450,255,window=month)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(month,4),activebackground="yellow")
            mycanvas.create_window(610,250,window=ser)
            
        if c=="Year":
            mycanvas.create_text(200,250,text="Enter the Year: ",font="Calibri 20 bold")
            year = Entry(root,font="Calibri 15 bold")
            mycanvas.create_window(440,255,window=year)
            ser = Button(root,text="Search",bg="brown",fg="white",font="Helvatica 15 bold",command=lambda: lock_goto(year,5),activebackground="yellow")
            mycanvas.create_window(600,250,window=ser)
            
        
    enter = Button(root,text="Continue",font="Calibri 15 bold",command=data,bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(400,180,window=enter)

    
    
    back3 = Button(root,text="Back",font="Calibri 15 bold",command=again,bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(50,30,window=back3)
    root.mainloop()


def delete_entry_overall():
    global root1,search_zero
    global bg3
    #def revisit_menu2():
    def again():
        try:
            search_zero.destroy()
            menu()
        except:
            pass
        
    try:
        root1.destroy()
    except:
        pass
    search_zero = Tk()
    search_zero.title("Account Management System")
    mycanvas = Canvas(search_zero,width=1920,height=1080)
    mycanvas.pack()
    #set background
    bg3 = PhotoImage(file="bg_2.png")
    mycanvas.create_image(0,0,image=bg3,anchor="nw")
    mycanvas.create_text(int(1920/2)-200,40,text="Account Management System",font="Helvatica 30 bold")

    #add other options
    button1 = Button(search_zero,text="Delete in Purchase Invoice",command=delete1,font="Calibri 30",bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(int(1920/2)-200,250,window=button1)

    button2 = Button(search_zero,text="Delete in Sales Invoice",command=delete2,font="Calibri 30",bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(int(1920/2)-200,550,window=button2)

    back3 = Button(search_zero,text="Back",font="Calibri 15 bold",command=again,bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(50,30,window=back3)
    search_zero.mainloop()




# DELETE Sales    

def delete_2(n,num,purchase,sheet):
    found_list = []
    global sheet1
    sheet1 = sheet
    found = False
    flag = []
    global c,row_values,submit
    if (num == 1):

        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    found = True
                    found_list.append(row_values)
                    flag.append("1")
            n = n+1
        
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Sales invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Sales invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Sales invoice number",text = "Sales invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)

            length = len(flag)
            if length == 1:
        
                def text():
                    submit.config(state=DISABLED)
                    inv.config(state=DISABLED)
                    global inv_number
                    try:
                        inv_number=int(inv.get())
                    except:
                        mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                    final_ask(sheet)

                    
                mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
                inv = Entry(root,font="Helvatica 15 bold")
                mycanvas.create_window(510,650,window=inv)
                submit = Button(root,text="Continue",command=text,font="Helvatica 15 bold",bg="brown",fg="white")
                mycanvas.create_window(700,650,window=submit)

        length = len(flag)   
        if (length > 1):
            
            def text2():
                submit.config(state=DISABLED)
                inv.config(state=DISABLED)
                global inv_number,c,row_values,sheet1
                sheet = sheet1
                try:
                    inv_number=int(inv.get())
                except:
                    mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                if submit['state']==DISABLED:
                    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                        for cell in row:
                            if str(cell.value) == str(inv_number):
                                sheet.cell(row=cell.row, column=cell.column)
                                c = cell.row
                                row_values = [cell.value for cell in sheet[c]]
                final_ask(sheet)

                
                   
            mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
            inv = Entry(root,font="Helvatica 15 bold")
            mycanvas.create_window(510,650,window=inv)
            
            submit = Button(root,text="Continue",command=text2,font="Helvatica 15 bold",bg="brown",fg="white")
            mycanvas.create_window(700,650,window=submit)

                

        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')
        
    elif (num == 2):

        for row in sheet.iter_rows(min_row=1, min_col=5, max_row=n, max_col=5):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    found = True
                    found_list.append(row_values)
                    flag.append("1")
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Sales invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Sales invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Sales invoice number",text = "Sales invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)

            length = len(flag)
            if length == 1:
        
                def text3():
                    submit.config(state=DISABLED)
                    inv.config(state=DISABLED)
                    global inv_number
                    try:
                        inv_number=int(inv.get())
                    except:
                        mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                    final_ask(sheet)

                    
                mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
                inv = Entry(root,font="Helvatica 15 bold")
                mycanvas.create_window(510,650,window=inv)
                submit = Button(root,text="Continue",command=text3,font="Helvatica 15 bold",bg="brown",fg="white")
                mycanvas.create_window(700,650,window=submit)

        length = len(flag)
        if (length > 1):

            def text4():
                submit.config(state=DISABLED)
                inv.config(state=DISABLED)
                global inv_number,c,row_values,sheet1
                sheet = sheet1
                try:
                    inv_number=int(inv.get())
                except:
                    mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                if submit['state']==DISABLED:
                    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                        for cell in row:
                            if str(cell.value) == str(inv_number):
                                sheet.cell(row=cell.row, column=cell.column)
                                c = cell.row
                                row_values = [cell.value for cell in sheet[c]]
                final_ask(sheet)                

            mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
            inv = Entry(root,font="Helvatica 15 bold")
            mycanvas.create_window(510,650,window=inv)
            
            submit = Button(root,text="Continue",command=text4,font="Helvatica 15 bold",bg="brown",fg="white")
            mycanvas.create_window(700,650,window=submit)


        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

            
    elif (num == 3):

        for row in sheet.iter_rows(min_row=1, min_col=3, max_row=n, max_col=3):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    found = True
                    found_list.append(row_values)
                    flag.append("1")
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Sales invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Sales invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Sales invoice number",text = "Sales invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)

            length = len(flag)
            if length == 1:
        
                def text5():
                    submit.config(state=DISABLED)
                    inv.config(state=DISABLED)
                    global inv_number
                    try:
                        inv_number=int(inv.get())
                    except:
                        mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                    final_ask(sheet)

                    
                mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
                inv = Entry(root,font="Helvatica 15 bold")
                mycanvas.create_window(510,650,window=inv)
                submit = Button(root,text="Continue",command=text5,font="Helvatica 15 bold",bg="brown",fg="white")
                mycanvas.create_window(700,650,window=submit)

        length = len(flag)
        if (length > 1):
            def text6():
                submit.config(state=DISABLED)
                inv.config(state=DISABLED)
                global inv_number,c,row_values,sheet1
                sheet = sheet1
                try:
                    inv_number=int(inv.get())
                except:
                    mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                if submit['state']==DISABLED:
                    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                        for cell in row:
                            if str(cell.value) == str(inv_number):
                                sheet.cell(row=cell.row, column=cell.column)
                                c = cell.row
                                row_values = [cell.value for cell in sheet[c]]
                final_ask(sheet)

            mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
            inv = Entry(root,font="Helvatica 15 bold")
            mycanvas.create_window(510,650,window=inv)
            
            submit = Button(root,text="Continue",command=text6,font="Helvatica 15 bold",bg="brown",fg="white")
            mycanvas.create_window(700,650,window=submit)

        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

    elif (num == 4):

        for row in sheet.iter_rows(min_row=1, min_col=2, max_row=n, max_col=2):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    found = True
                    found_list.append(row_values)
                    flag.append("1")
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Sales invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Sales invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Sales invoice number",text = "Sales invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)

            length = len(flag)
            if length == 1:
        
                def text7():
                    submit.config(state=DISABLED)
                    inv.config(state=DISABLED)
                    global inv_number
                    try:
                        inv_number=int(inv.get())
                    except:
                        mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                    final_ask(sheet)

                    
                mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
                inv = Entry(root,font="Helvatica 15 bold")
                mycanvas.create_window(510,650,window=inv)
                submit = Button(root,text="Continue",command=text7,font="Helvatica 15 bold",bg="brown",fg="white")
                mycanvas.create_window(700,650,window=submit)

        length = len(flag)
        if (length > 1):
            
            def text8():
                submit.config(state=DISABLED)
                inv.config(state=DISABLED)
                global inv_number,c,row_values,sheet1
                sheet = sheet1
                try:
                    inv_number=int(inv.get())
                except:
                    mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                if submit['state']==DISABLED:
                    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                        for cell in row:
                            if str(cell.value) == str(inv_number):
                                sheet.cell(row=cell.row, column=cell.column)
                                c = cell.row
                                row_values = [cell.value for cell in sheet[c]]
                final_ask(sheet)
            mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
            inv = Entry(root,font="Helvatica 15 bold")
            mycanvas.create_window(510,650,window=inv)
            submit = Button(root,text="Continue",command=text8,font="Helvatica 15 bold",bg="brown",fg="white")
            mycanvas.create_window(700,650,window=submit)
                    
        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

    elif (num == 5):

        for row in sheet.iter_rows(min_row=1, min_col=4, max_row=n, max_col=4):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    found_list.append(row_values)
                    flag.append("1")
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Sales invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Sales invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Sales invoice number",text = "Sales invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)

            length = len(flag)
            if length == 1:
        
                def text9():
                    submit.config(state=DISABLED)
                    inv.config(state=DISABLED)
                    global inv_number
                    try:
                        inv_number=int(inv.get())
                    except:
                        mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                    final_ask(sheet)

                    
                mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
                inv = Entry(root,font="Helvatica 15 bold")
                mycanvas.create_window(510,650,window=inv)
                submit = Button(root,text="Continue",command=text9,font="Helvatica 15 bold",bg="brown",fg="white")
                mycanvas.create_window(700,650,window=submit)

        length = len(flag)
        if (length > 1):
            def text10():
                submit.config(state=DISABLED)
                inv.config(state=DISABLED)
                global inv_number,c,row_values,sheet1
                sheet = sheet1
                try:
                    inv_number=int(inv.get())
                except:
                    mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                if submit['state']==DISABLED:
                    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                        for cell in row:
                            if str(cell.value) == str(inv_number):
                                sheet.cell(row=cell.row, column=cell.column)
                                c = cell.row
                                row_values = [cell.value for cell in sheet[c]]
                    final_ask(sheet)
            mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
            inv = Entry(root,font="Helvatica 15 bold")
            mycanvas.create_window(510,650,window=inv)
            submit = Button(root,text="Continue",command=text10,font="Helvatica 15 bold",bg="brown",fg="white")
            mycanvas.create_window(700,650,window=submit)

        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')


# delete purchase

def delete(n,num,purchase,sheet):
    found_list = []
    global sheet1
    sheet1 = sheet
    found = False
    flag = []
    global c,row_values,submit
    if (num == 1):

        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    found = True
                    found_list.append(row_values)
                    flag.append("1")
            n = n+1
        
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Purchase invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Purchase invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Purchase invoice number",text = "Purchase invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)

            length = len(flag)
            if length == 1:
        
                def text():
                    submit.config(state=DISABLED)
                    inv.config(state=DISABLED)
                    global inv_number
                    try:
                        inv_number=int(inv.get())
                    except:
                        mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                    final_ask(sheet)

                    
                mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
                inv = Entry(root,font="Helvatica 15 bold")
                mycanvas.create_window(510,650,window=inv)
                submit = Button(root,text="Continue",command=text,font="Helvatica 15 bold",bg="brown",fg="white")
                mycanvas.create_window(700,650,window=submit)

        length = len(flag)   
        if (length > 1):
            
            def text2():
                submit.config(state=DISABLED)
                inv.config(state=DISABLED)
                global inv_number,c,row_values,sheet1
                sheet = sheet1
                try:
                    inv_number=int(inv.get())
                except:
                    mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                if submit['state']==DISABLED:
                    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                        for cell in row:
                            if str(cell.value) == str(inv_number):
                                sheet.cell(row=cell.row, column=cell.column)
                                c = cell.row
                                row_values = [cell.value for cell in sheet[c]]
                final_ask(sheet)

                
                   
            mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
            inv = Entry(root,font="Helvatica 15 bold")
            mycanvas.create_window(510,650,window=inv)
            
            submit = Button(root,text="Continue",command=text2,font="Helvatica 15 bold",bg="brown",fg="white")
            mycanvas.create_window(700,650,window=submit)

                

        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')
        
    elif (num == 2):

        for row in sheet.iter_rows(min_row=1, min_col=5, max_row=n, max_col=5):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    found = True
                    found_list.append(row_values)
                    flag.append("1")
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Purchase invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Purchase invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Purchase invoice number",text = "Purchase invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)

            length = len(flag)
            if length == 1:
        
                def text3():
                    submit.config(state=DISABLED)
                    inv.config(state=DISABLED)
                    global inv_number
                    try:
                        inv_number=int(inv.get())
                    except:
                        mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                    final_ask(sheet)

                    
                mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
                inv = Entry(root,font="Helvatica 15 bold")
                mycanvas.create_window(510,650,window=inv)
                submit = Button(root,text="Continue",command=text3,font="Helvatica 15 bold",bg="brown",fg="white")
                mycanvas.create_window(700,650,window=submit)

        length = len(flag)
        if (length > 1):

            def text4():
                submit.config(state=DISABLED)
                inv.config(state=DISABLED)
                global inv_number,c,row_values,sheet1
                sheet = sheet1
                try:
                    inv_number=int(inv.get())
                except:
                    mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                if submit['state']==DISABLED:
                    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                        for cell in row:
                            if str(cell.value) == str(inv_number):
                                sheet.cell(row=cell.row, column=cell.column)
                                c = cell.row
                                row_values = [cell.value for cell in sheet[c]]
                final_ask(sheet)                

            mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
            inv = Entry(root,font="Helvatica 15 bold")
            mycanvas.create_window(510,650,window=inv)
            
            submit = Button(root,text="Continue",command=text4,font="Helvatica 15 bold",bg="brown",fg="white")
            mycanvas.create_window(700,650,window=submit)


        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

            
    elif (num == 3):

        for row in sheet.iter_rows(min_row=1, min_col=3, max_row=n, max_col=3):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    found = True
                    found_list.append(row_values)
                    flag.append("1")
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Purchase invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Purchase invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Purchase invoice number",text = "Purchase invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)

            length = len(flag)
            if length == 1:
        
                def text5():
                    submit.config(state=DISABLED)
                    inv.config(state=DISABLED)
                    global inv_number
                    try:
                        inv_number=int(inv.get())
                    except:
                        mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                    final_ask(sheet)

                    
                mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
                inv = Entry(root,font="Helvatica 15 bold")
                mycanvas.create_window(510,650,window=inv)
                submit = Button(root,text="Continue",command=text5,font="Helvatica 15 bold",bg="brown",fg="white")
                mycanvas.create_window(700,650,window=submit)

        length = len(flag)
        if (length > 1):
            def text6():
                submit.config(state=DISABLED)
                inv.config(state=DISABLED)
                global inv_number,c,row_values,sheet1
                sheet = sheet1
                try:
                    inv_number=int(inv.get())
                except:
                    mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                if submit['state']==DISABLED:
                    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                        for cell in row:
                            if str(cell.value) == str(inv_number):
                                sheet.cell(row=cell.row, column=cell.column)
                                c = cell.row
                                row_values = [cell.value for cell in sheet[c]]
                final_ask(sheet)

            mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
            inv = Entry(root,font="Helvatica 15 bold")
            mycanvas.create_window(510,650,window=inv)
            
            submit = Button(root,text="Continue",command=text6,font="Helvatica 15 bold",bg="brown",fg="white")
            mycanvas.create_window(700,650,window=submit)

        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

    elif (num == 4):

        for row in sheet.iter_rows(min_row=1, min_col=2, max_row=n, max_col=2):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    found = True
                    found_list.append(row_values)
                    flag.append("1")
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Purchase invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Purchase invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Purchase invoice number",text = "Purchase invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)

            length = len(flag)
            if length == 1:
        
                def text7():
                    submit.config(state=DISABLED)
                    inv.config(state=DISABLED)
                    global inv_number
                    try:
                        inv_number=int(inv.get())
                    except:
                        mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                    final_ask(sheet)

                    
                mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
                inv = Entry(root,font="Helvatica 15 bold")
                mycanvas.create_window(510,650,window=inv)
                submit = Button(root,text="Continue",command=text7,font="Helvatica 15 bold",bg="brown",fg="white")
                mycanvas.create_window(700,650,window=submit)

        length = len(flag)
        if (length > 1):
            
            def text8():
                submit.config(state=DISABLED)
                inv.config(state=DISABLED)
                global inv_number,c,row_values,sheet1
                sheet = sheet1
                try:
                    inv_number=int(inv.get())
                except:
                    mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                if submit['state']==DISABLED:
                    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                        for cell in row:
                            if str(cell.value) == str(inv_number):
                                sheet.cell(row=cell.row, column=cell.column)
                                c = cell.row
                                row_values = [cell.value for cell in sheet[c]]
                final_ask(sheet)
            mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
            inv = Entry(root,font="Helvatica 15 bold")
            mycanvas.create_window(510,650,window=inv)
            submit = Button(root,text="Continue",command=text8,font="Helvatica 15 bold",bg="brown",fg="white")
            mycanvas.create_window(700,650,window=submit)
                    
        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')

    elif (num == 5):

        for row in sheet.iter_rows(min_row=1, min_col=4, max_row=n, max_col=4):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    found_list.append(row_values)
                    flag.append("1")
            n = n+1
        if found:
            tree = ttk.Treeview(root)
            tree['columns']=['Purchase invoice number','MM','DD','YYYY','Party name','Party address','GST_num','PAN_num','Mode of Payment','Bill_num','Product','Rate','Quantity','Total','GST','Total Bill']
            tree.column("#0",width=80,stretch=NO)
            tree.column("Purchase invoice number",width=80)
            tree.column("MM",width=80)
            tree.column("DD",width=80)
            tree.column("YYYY",width=80)
            tree.column("Party name",width=80)
            tree.column("Party address",width=80)
            tree.column("GST_num",width=80)
            tree.column("PAN_num",width=80)
            tree.column("Mode of Payment",width=80)
            tree.column("Bill_num",width=80)
            tree.column('Product',width=80)
            tree.column("Rate",width=80)
            tree.column("Quantity",width=80)
            tree.column("Total",width=80)
            tree.column("GST",width=80)
            tree.column("Total Bill",width=80)
            tree.heading("#0",text = "Sr.no.")
            tree.heading("Purchase invoice number",text = "Purchase invoice number")
            tree.heading("MM",text = "MM")
            tree.heading("DD",text = "DD")
            tree.heading("YYYY",text = "YYYY")
            tree.heading("Party name",text = "Party name")
            tree.heading("Party address",text = "Party address")
            tree.heading("GST_num",text = "GST_num")
            tree.heading("PAN_num",text = "PAN_num")
            tree.heading("Mode of Payment",text = "Mode of Payment")
            tree.heading("Bill_num",text = "Bill_num")
            tree.heading('Product',text = "Product")
            tree.heading("Rate",text = "Rate")
            tree.heading("Quantity",text = "Quantity")
            tree.heading("Total",text = "Total")
            tree.heading("GST",text = "GST")
            tree.heading("Total Bill",text = "Total Bill")
            for i in range(0,len(found_list)):
                tree.insert(parent='',index='end',iid=i,text=str(i+1),values=found_list[i])

            mycanvas.create_window(750,450,window=tree)

            length = len(flag)
            if length == 1:
        
                def text9():
                    submit.config(state=DISABLED)
                    inv.config(state=DISABLED)
                    global inv_number
                    try:
                        inv_number=int(inv.get())
                    except:
                        mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                    final_ask(sheet)

                    
                mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
                inv = Entry(root,font="Helvatica 15 bold")
                mycanvas.create_window(510,650,window=inv)
                submit = Button(root,text="Continue",command=text9,font="Helvatica 15 bold",bg="brown",fg="white")
                mycanvas.create_window(700,650,window=submit)

        length = len(flag)
        if (length > 1):
            def text10():
                submit.config(state=DISABLED)
                inv.config(state=DISABLED)
                global inv_number,c,row_values,sheet1
                sheet = sheet1
                try:
                    inv_number=int(inv.get())
                except:
                    mycanvas.create_text(200,800,text="Please enter valid integer invoice number.",fill='red',font="Helvatica 15 bold")
                if submit['state']==DISABLED:
                    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                        for cell in row:
                            if str(cell.value) == str(inv_number):
                                sheet.cell(row=cell.row, column=cell.column)
                                c = cell.row
                                row_values = [cell.value for cell in sheet[c]]
                    final_ask(sheet)
            mycanvas.create_text(200,650,text="Enter the invoice number to delete:",font="Helvatica 15 bold")
            inv = Entry(root,font="Helvatica 15 bold")
            mycanvas.create_window(510,650,window=inv)
            submit = Button(root,text="Continue",command=text10,font="Helvatica 15 bold",bg="brown",fg="white")
            mycanvas.create_window(700,650,window=submit)

        if not found:
            mycanvas.create_text(int(1920/2)-200,350,text="No entry found!",font="Helvatica 40 bold",fill='red')


def final_ask(sheet):
    def no():
        top.destroy()
        label = Label(root,text="Are you sure that you want to delete this row?",font="Helvatica 15 bold")
        mycanvas.create(200,750,text="Thank you for using the program.",fill="cyan",font="Helvatica 20 bold")
    def yeah():
        global c
        global row_values
        try:
            top.destroy()
        except:
            pass
        global cd
        cd = 1
        if (cd == 1):
                           
            sheet.delete_rows(idx=c,amount=1)
            if sheet==wb["purchase invoice"]:
                cb["A1"].value=cb["A1"].value+row_values[-1]
            if sheet==wb["sales invoice"]:
                cb["A1"].value=cb["A1"].value-row_values[-1]
        mycanvas.create_text(int(1920/2)-200,700,text="Entry deleted.",font="Helvatica 40 bold",fill='red')
        wb.save("database_project.xlsx")
        wbcb.save("current_balance.xlsx")
    
    top = Toplevel()
    top.title("Account Management System")
    label = Label(top,text="Are you sure that you want to delete this row?",font="Helvatica 15 bold")
    label.grid(row=0,column=0,columnspan=2)
    but1 = Button(top,text="Yes",command=yeah,fg="white",bg="black",activebackground="red",font="Helvatica 15")
    but2 = Button(top,text="No",command=no,fg="white",bg="cyan",activebackground="yellow",font="Helvatica 15 bold")
    but1.grid(row=1,column=0)
    but2.grid(row=1,column=1)



# update
def update_data(n):
    try:
        root1.destroy()
    except:
        pass
    def updatepro(row_values,c,sheet):
        print(len(row_values))
        wbcb=openpyxl.load_workbook("current_balance.xlsx")
        cb=wbcb["Sheet1"]
        duplicate=[]
        for ele in row_values:
            duplicate.append(ele)
        cb_old=duplicate[15]
        cb["A1"].value=cb["A1"].value-cb_old
        choose = input("\nNow enter the column name which you want to update in this row: ")

        print("="*43)

    
        if ((choose=='purchase invoice number') or (choose=='sales invoice number') or (choose=='party name') or (choose=='party adress') or (choose=='gst number') or (choose=='mode of payment')
        or (choose=='product')):        
            replace = input("Enter the new value to update: ")
            for row in sheet.iter_rows(min_row=1, min_col=1, max_col=17, max_row=1):
                for cell in row:
                    if cell.value == choose:
                        sheet.cell(row=c, column=cell.column)
                        s = cell.column
                        #c = cell.row
                        print(s,",",c)
                        sheet.cell(row=c, column=s).value = replace
                        row_values = [cell.value for cell in sheet[c]]
                        print("\nUpdated entries: ", row_values)

        elif ((choose=='MM') or (choose=='DD') or (choose=='YYYY') or (choose=='pan number') or (choose=='bill number')):
            replace = int(input("Enter the new value to update: "))
            for row in sheet.iter_rows(min_row=1, min_col=1, max_col=17, max_row=1):
                for cell in row:
                    if cell.value == choose:
                            sheet.cell(row=c, column=cell.column)
                            s = cell.column
                            #c = cell.row
                            sheet.cell(row=c, column=s).value = replace
                            row_values = [cell.value for cell in sheet[c]]
                            print("\nUpdated entries: ", row_values)
        elif ((choose=='quantity') or (choose=='rate') or (choose=='gst')):
            replace = int(input("Enter the new value to update: "))
            cb["A1"].value=cb["A1"].value-cb_old
            for row in sheet.iter_rows(min_row=1, min_col=1, max_col=17, max_row=1):
                for cell in row:
                    if cell.value == choose:
                        sheet.cell(row=c, column=cell.column)
                        s = cell.column
                        #c = cell.row
                        sheet.cell(row=c, column=s).value = replace
                        row_values = [cell.value for cell in sheet[c]]
 
            print(cb["A1"].value)
            
            row_values[13]=row_values[12]*row_values[11]
            row_values[15]=((row_values[14]/100)*(row_values[13]))+row_values[13]
            

            print("\nUpdated entries: \n", row_values)

            sheet.cell(row=c,column=17).value=row_values[15]
            sheet.cell(row=c,column=16).value=row_values[14]
            sheet.cell(row=c,column=15).value=row_values[13]
            
        else:
            print("enter correct column name pls")
            updatepro(row_values,c,sheet)
        cb["A1"].value=cb["A1"].value+row_values[15]
        wb.save("database_project.xlsx")
        wbcb.save("current_balance.xlsx")

        
    print("To update in purchase invoice sheet press 1\nTo update in sales invoice sheet press 2\n")
    shch=input("enter choice here :")
    if shch=='1':
        sheet=wb["purchase invoice"]
    elif shch=='2':
        sheet=wb["sales invoice"]
    elif shch=='3':
        print("you have exited the function")
        #break
    else:
        print("kindly enter 1,2 or 3 only")

    print("="*20)
    
    print("====MENU===\n1.Search by invoice number.\n2.Search by name.\n3.Search by date.\n4.Search by month.\n5.Search by year.\n")
    num = int(input("\nEnter any number from the menu: "))
    print("="*33)
    found = False

    if (num == 1):
        flag = []

        purchase = input("Enter the number you want to search: ")
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    flag.append("1")

            n = n+1

        length = len(flag)
        if (length==1):
            updatepro(row_values,c,sheet)
        if (length > 1):
            inv_number = int(
                input("\nEnter invoice number of the entry you want to update: "))
            for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                for cell in row:
                    if str(cell.value) == str(inv_number):
                        sheet.cell(row=cell.row, column=cell.column)
                        c = cell.row
                        row_values = [cell.value for cell in sheet[c]]
                        print("\nFINDING...")
                        print("YOUR ENTRIES ARE:")
                        print("\n", row_values)
            updatepro(row_values,c,sheet)
        if not found:
            print("No entry found")
        

    elif (num == 2):
        flag = []

        purchase = input("Enter the name you want to search: ")
        for row in sheet.iter_rows(min_row=1, min_col=5, max_row=n, max_col=5):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    flag.append("1")
            n = n+1
        print(c)
        length = len(flag)
        if (length==1):
            updatepro(row_values,c,sheet)
        if (length > 1):
            inv_number = int(
                input("\nEnter invoice number of the entry you want to update: "))
            for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                for cell in row:
                    if str(cell.value) == str(inv_number):
                        sheet.cell(row=cell.row, column=cell.column)
                        c = cell.row
                        row_values = [cell.value for cell in sheet[c]]
                        print("\nFINDING...")
                        print("YOUR ENTRIES ARE:")
                        print("\n", row_values)
            updatepro(row_values,c,sheet)
        if not found:
            print("No entry found")

    elif (num == 3):
        flag = []

        purchase = input("Enter the number you want to search: ")
        for row in sheet.iter_rows(min_row=1, min_col=3, max_row=n, max_col=3):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    flag.append("1")
            n = n+1
        length = len(flag)
        if (length==1):
            updatepro(row_values,c,sheet)
        if (length > 1):
            inv_number = int(
                input("\nEnter invoice number of the entry you want to update: "))
            for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                for cell in row:
                    if str(cell.value) == str(inv_number):
                        sheet.cell(row=cell.row, column=cell.column)
                        c = cell.row
                        row_values = [cell.value for cell in sheet[c]]
                        print("\nFINDING...")
                        print("YOUR ENTRIES ARE:")
                        print("\n", row_values)
            updatepro(row_values,c,sheet)
        if not found:
            print("No entry found")

    elif (num == 4):
        flag = []

        purchase = input("Enter the number you want to search: ")
        for row in sheet.iter_rows(min_row=1, min_col=2, max_row=n, max_col=2):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    flag.append("1")
            n = n+1
        length = len(flag)
        if (length==1):
            updatepro(row_values,c,sheet)
        if (length > 1):
            inv_number = int(
                input("\nEnter invoice number of the entry you want to update: "))
            for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                for cell in row:
                    if str(cell.value) == str(inv_number):
                        sheet.cell(row=cell.row, column=cell.column)
                        c = cell.row
                        row_values = [cell.value for cell in sheet[c]]
                        print("\nFINDING...")
                        print("YOUR ENTRIES ARE:")
                        print("\n", row_values)
            updatepro(row_values,c,sheet)
        if not found:
            print("No entry found")
            
    elif (num == 5):
        flag = []

        purchase = input("Enter the number you want to search: ")
        for row in sheet.iter_rows(min_row=1, min_col=4, max_row=n, max_col=4):
            for cell in row:
                if str(cell.value) == str(purchase):
                    sheet.cell(row=cell.row, column=cell.column)
                    c = cell.row
                    row_values = [cell.value for cell in sheet[c]]
                    print("\nFINDING...")
                    print("YOUR ENTRIES ARE:")
                    print("\n", row_values)
                    found = True
                    flag.append("1")
            n = n+1
            
        length = len(flag)
        if (length==1):
            updatepro(row_values,c,sheet)
        elif (length > 1):
            inv_number = int(
                input("\nEnter invoice number of the entry you want to update: "))
            for row in sheet.iter_rows(min_row=1, min_col=1, max_row=n, max_col=1):
                for cell in row:
                    if str(cell.value) == str(inv_number):
                        sheet.cell(row=cell.row, column=cell.column)
                        c = cell.row
                        row_values = [cell.value for cell in sheet[c]]
                        print("\n", row_values)
            updatepro(row_values,c,sheet)


        if not found:
            print("No entry found")


            
    else:
        print("Enter a number from the list only")


    wb.save("database_project.xlsx")
    wbcb.save("current_balance.xlsx")
    print("----All changes saved----")
    menu()


#***************************************************************************************
# ****************************************graph
import matplotlib.pyplot as plt
global file
file = pd.read_excel('database_project.xlsx')


def graph_plot():
    global bg10,file
    def back_me():
        try:
            root3.destroy()
        except:
            pass
        menu()

    try:
        root1.destroy()
    except:
        pass
    root3 = Tk()
    root3.title("Account Management System")
    mycanvas = Canvas(root3,width=1920,height=1080)
    mycanvas.pack()
    bg10 = PhotoImage(file="bg_2.png")
    mycanvas.create_image(0,0,image=bg10,anchor="nw")
    mycanvas.create_text(int(1920/2)-200,40,text="Account Management System",font="Helvatica 30 bold")
    mycanvas.create_text(int(1920/2)-180,100,text="Graphs",font="Calibri 25 bold")

    button1 = Button(root3,text="Party Name vs Total Bill",command=plot1,font="Calibri 30",bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(300,200,window=button1)

    button2 = Button(root3,text="Total Bill vs Mode of Payment",command=plot2,font="Calibri 30",bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(300,450,window=button2)

    button3 = Button(root3,text="Product vs Total Bill",command=plot3,font="Calibri 30",bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(1200,200,window=button3)

    button4 = Button(root3,text="Year vs Product",command=plot4,font="Calibri 30",bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(1200,450,window=button4)

    button5 = Button(root3,text="Address vs Total Bill",command=plot5,font="Calibri 30",bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(int(1920/2)-150,650,window=button5)

    button6 = Button(root3,text="Back",command=back_me,font="Calibri 15",bg="brown",fg="white",activebackground="yellow")
    mycanvas.create_window(100,50,window=button6)
    

    root3.mainloop()
#plot 1

def plot1():
    def back_graph():
        try:
            top.destroy()
        except:
            pass
    
    def dis_year():
        file = pd.read_excel('database_project.xlsx')
        def plot():
            year_str = entry.get()
            if year_str.isdigit():
                year_value = int(year_str)
                allow = 1
            else:
                allow = 0
                req = Label(top,text="Please enter a valid year (yyyy-integers)",font="Helvatica 15 bold")
                canvas.create_window(200,325,window=req)
            if allow == 1:
                try:
                    top.destroy()
                except:
                    pass
                year_column = 'YYYY'
                year_data = file[file[year_column] == year_value]
                category_column = 'party name'
                value_column = 'total bill'
                categories = year_data[category_column].unique()
                values = year_data[value_column].values
                fig, ax = plt.subplots()
                ax.bar(categories, values)
                ax.set_xlabel("Party Name")
                ax.set_ylabel("Total Bill")
                ax.set_title(f'Bar chart of Total Bill by Party Name for {year_value}')
                plt.show()
                    
        but1.config(state=DISABLED)
        but2.config(state=DISABLED)
        ask = Label(top,text="Enter the year: ",font="Helvatica 15 bold")
        entry = Entry(top,font="Helvatica 15")
        sub = Button(top,font="Helvatica 15 bold",command=plot,text="Submit",bg="light blue",fg="white")
        canvas.create_window(100,250,window=ask)
        canvas.create_window(300,250,window=entry)
        canvas.create_window(500,250,window=sub)
    

    def all_year():
        try:
            top.destroy()
        except:
            pass
        x_axis = file['party name']
        y_axis = file['total bill']
        plt.bar(x_axis, y_axis, width=0.5)
        plt.xlabel("Party Name")
        plt.ylabel("Total Bill")
        plt.show()

        
    top=Toplevel()
    canvas = Canvas(top,height=500,width=1200)
    canvas.pack()
    label = Label(top,text="Display graph: Kindly choose from the options below",font="Helvatica 15")
    but1 = Button(top,text="Display from a particular year",command=dis_year,font="Helvatica 20 bold",bg="cyan",fg="white",activebackground="yellow")
    but2 = Button(top,text="Display through all years",command=all_year,font="Helvatica 20 bold",bg="green",fg="white",activebackground="yellow")
    but3 = Button(top,text="Back",command=back_graph,font="Helvatica 20 bold",bg="brown",fg="white",activebackground="yellow")

    canvas.create_window(250,40,window=label)
    canvas.create_window(250,150,window=but1)
    canvas.create_window(700,150,window=but2)
    canvas.create_window(1100,150,window=but3)
    top.mainloop()

# plot 2
def plot2():
    def back_graph():
        try:
            top.destroy()
        except:
            pass
    
    def dis_year():
        file = pd.read_excel('database_project.xlsx')
        def plot():
            year_str = entry.get()
            if year_str.isdigit():
                year_value = int(year_str)
                allow = 1
            else:
                allow = 0
                req = Label(top,text="Please enter a valid year (yyyy-integers)",font="Helvatica 15 bold")
                canvas.create_window(200,325,window=req)
            if allow == 1:
                try:
                    top.destroy()
                except:
                    pass
                year_column = 'YYYY'
                year_value = int(input("Enter a year: "))
                year_data = file[file[year_column] == year_value]
                category_column = 'mode of payment'
                value_column = 'total bill'
                categories = year_data[category_column].unique()
                values = year_data[value_column].values
                fig, ax = plt.subplots()
                ax.bar(categories, values)
                ax.set_xlabel("Mode of Payment")
                ax.set_ylabel("Total Bill")
                ax.set_title(f'Bar chart of Total Bill by Mode of Payment for {year_value}')
                plt.show()
                    
        but1.config(state=DISABLED)
        but2.config(state=DISABLED)
        ask = Label(top,text="Enter the year: ",font="Helvatica 15 bold")
        entry = Entry(top,font="Helvatica 15")
        sub = Button(top,font="Helvatica 15 bold",command=plot,text="Submit",bg="light blue",fg="white")
        canvas.create_window(100,250,window=ask)
        canvas.create_window(300,250,window=entry)
        canvas.create_window(500,250,window=sub)
    

    def all_year():
        try:
            top.destroy()
        except:
            pass
        y_axis = file['total bill']
        x_axis = file['mode of payment']
        plt.bar(x_axis, y_axis, width=0.5)
        plt.xlabel("Mode of Payment")
        plt.ylabel("Total Bill")
        plt.show()

        
    top=Toplevel()
    canvas = Canvas(top,height=500,width=1200)
    canvas.pack()
    label = Label(top,text="Display graph: Kindly choose from the options below",font="Helvatica 15")
    but1 = Button(top,text="Display from a particular year",command=dis_year,font="Helvatica 20 bold",bg="cyan",fg="white",activebackground="yellow")
    but2 = Button(top,text="Display through all years",command=all_year,font="Helvatica 20 bold",bg="green",fg="white",activebackground="yellow")
    but3 = Button(top,text="Back",command=back_graph,font="Helvatica 20 bold",bg="brown",fg="white",activebackground="yellow")

    canvas.create_window(250,40,window=label)
    canvas.create_window(250,150,window=but1)
    canvas.create_window(700,150,window=but2)
    canvas.create_window(1100,150,window=but3)
    top.mainloop()


# graph 3

def plot3():
    def back_graph():
        try:
            top.destroy()
        except:
            pass
    
    def dis_year():
        file = pd.read_excel('database_project.xlsx')
        def plot():
            year_str = entry.get()
            if year_str.isdigit():
                year_value = int(year_str)
                allow = 1
            else:
                allow = 0
                req = Label(top,text="Please enter a valid year (yyyy-integers)",font="Helvatica 15 bold")
                canvas.create_window(200,325,window=req)
            if allow == 1:
                try:
                    top.destroy()
                except:
                    pass
                year_column = 'YYYY'
                year_value = int(input("Enter a year: "))
                year_data = file[file[year_column] == year_value]
                category_column = 'product'
                value_column = 'total bill'
                categories = year_data[category_column].unique()
                values = year_data[value_column].values
                fig, ax = plt.subplots()
                ax.bar(categories, values)
                ax.set_xlabel("Product")
                ax.set_ylabel("Total Bill")
                ax.set_title(f'Bar chart of Total Bill by Product for {year_value}')
                plt.show()
                    
        but1.config(state=DISABLED)
        but2.config(state=DISABLED)
        ask = Label(top,text="Enter the year: ",font="Helvatica 15 bold")
        entry = Entry(top,font="Helvatica 15")
        sub = Button(top,font="Helvatica 15 bold",command=plot,text="Submit",bg="light blue",fg="white")
        canvas.create_window(100,250,window=ask)
        canvas.create_window(300,250,window=entry)
        canvas.create_window(500,250,window=sub)
    

    def all_year():
        try:
            top.destroy()
        except:
            pass
        x_axis = file['product']
        y_axis = file['total bill']
        plt.bar(x_axis, y_axis, width=0.5)
        plt.xlabel("Product")
        plt.ylabel("Total Bill")
        plt.show()

        
    top=Toplevel()
    canvas = Canvas(top,height=500,width=1200)
    canvas.pack()
    label = Label(top,text="Display graph: Kindly choose from the options below",font="Helvatica 15")
    but1 = Button(top,text="Display from a particular year",command=dis_year,font="Helvatica 20 bold",bg="cyan",fg="white",activebackground="yellow")
    but2 = Button(top,text="Display through all years",command=all_year,font="Helvatica 20 bold",bg="green",fg="white",activebackground="yellow")
    but3 = Button(top,text="Back",command=back_graph,font="Helvatica 20 bold",bg="brown",fg="white",activebackground="yellow")

    canvas.create_window(250,40,window=label)
    canvas.create_window(250,150,window=but1)
    canvas.create_window(700,150,window=but2)
    canvas.create_window(1100,150,window=but3)
    top.mainloop()


# plot 4***********************************

def plot4():
    def back_graph():
        try:
            top.destroy()
        except:
            pass
    
    def dis_year():
        file = pd.read_excel('database_project.xlsx')
        def plot():
            year_str = entry.get()
            if year_str.isdigit():
                year_value = int(year_str)
                allow = 1
            else:
                allow = 0
                req = Label(top,text="Please enter a valid year (yyyy-integers)",font="Helvatica 15 bold")
                canvas.create_window(200,325,window=req)
            if allow == 1:
                try:
                    top.destroy()
                except:
                    pass
                year_column = 'YYYY'
                year_value = int(input("Enter a year: "))
                year_data = file[file[year_column] == year_value]
                category_column = 'product'
                value_column = 'YYYY'
                categories = year_data[category_column].unique()
                values = year_data[value_column].values
                fig, ax = plt.subplots()
                ax.bar(categories, values)
                ax.set_xlabel("Product")
                ax.set_ylabel("Year")
                ax.set_title(f'Bar chart of Year by Product for {year_value}')
                plt.show()
                    
        but1.config(state=DISABLED)
        but2.config(state=DISABLED)
        ask = Label(top,text="Enter the year: ",font="Helvatica 15 bold")
        entry = Entry(top,font="Helvatica 15")
        sub = Button(top,font="Helvatica 15 bold",command=plot,text="Submit",bg="light blue",fg="white")
        canvas.create_window(100,250,window=ask)
        canvas.create_window(300,250,window=entry)
        canvas.create_window(500,250,window=sub)
    

    def all_year():
        try:
            top.destroy()
        except:
            pass
        x_axis = file['YYYY']
        y_axis = file['product']
        plt.bar(x_axis, y_axis, width=0.5)
        plt.xlabel("Year")
        plt.ylabel("Product")
        plt.show()

        
    top=Toplevel()
    canvas = Canvas(top,height=500,width=1200)
    canvas.pack()
    label = Label(top,text="Display graph: Kindly choose from the options below",font="Helvatica 15")
    but1 = Button(top,text="Display from a particular year",command=dis_year,font="Helvatica 20 bold",bg="cyan",fg="white",activebackground="yellow")
    but2 = Button(top,text="Display through all years",command=all_year,font="Helvatica 20 bold",bg="green",fg="white",activebackground="yellow")
    but3 = Button(top,text="Back",command=back_graph,font="Helvatica 20 bold",bg="brown",fg="white",activebackground="yellow")

    canvas.create_window(250,40,window=label)
    canvas.create_window(250,150,window=but1)
    canvas.create_window(700,150,window=but2)
    canvas.create_window(1100,150,window=but3)
    top.mainloop()


# plot 5********************

def plot5():
    def back_graph():
        try:
            top.destroy()
        except:
            pass
    
    def dis_year():
        file = pd.read_excel('database_project.xlsx')
        def plot():
            year_str = entry.get()
            if year_str.isdigit():
                year_value = int(year_str)
                allow = 1
            else:
                allow = 0
                req = Label(top,text="Please enter a valid year (yyyy-integers)",font="Helvatica 15 bold")
                canvas.create_window(200,325,window=req)
            if allow == 1:
                try:
                    top.destroy()
                except:
                    pass
                year_column = 'YYYY'
                year_value = int(input("Enter a year: "))
                year_data = file[file[year_column] == year_value]
                category_column = 'party address'
                value_column = 'total bill'
                categories = year_data[category_column].unique()
                values = year_data[value_column].values
                fig, ax = plt.subplots()
                ax.bar(categories, values)
                ax.set_xlabel("Party Address")
                ax.set_ylabel("Total Bill")
                ax.set_title(f'Bar chart of Total Bill by Party Address for {year_value}')
                plt.show()
                
                    
        but1.config(state=DISABLED)
        but2.config(state=DISABLED)
        ask = Label(top,text="Enter the year: ",font="Helvatica 15 bold")
        entry = Entry(top,font="Helvatica 15")
        sub = Button(top,font="Helvatica 15 bold",command=plot,text="Submit",bg="light blue",fg="white")
        canvas.create_window(100,250,window=ask)
        canvas.create_window(300,250,window=entry)
        canvas.create_window(500,250,window=sub)
    

    def all_year():
        try:
            top.destroy()
        except:
            pass
        x_axis = file['party adress']
        y_axis = file['total bill']
        plt.bar(x_axis, y_axis, width=0.5)
        plt.xlabel("Party Address")
        plt.ylabel("Total Bill")
        plt.show()

        
    top=Toplevel()
    canvas = Canvas(top,height=500,width=1200)
    canvas.pack()
    label = Label(top,text="Display graph: Kindly choose from the options below",font="Helvatica 15")
    but1 = Button(top,text="Display from a particular year",command=dis_year,font="Helvatica 20 bold",bg="cyan",fg="white",activebackground="yellow")
    but2 = Button(top,text="Display through all years",command=all_year,font="Helvatica 20 bold",bg="green",fg="white",activebackground="yellow")
    but3 = Button(top,text="Back",command=back_graph,font="Helvatica 20 bold",bg="brown",fg="white",activebackground="yellow")

    canvas.create_window(250,40,window=label)
    canvas.create_window(250,150,window=but1)
    canvas.create_window(700,150,window=but2)
    canvas.create_window(1100,150,window=but3)
    top.mainloop()



root.mainloop()
